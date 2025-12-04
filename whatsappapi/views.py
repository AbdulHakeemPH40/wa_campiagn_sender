"""
WASender Integration Views
Handles WhatsApp session management and campaign sending
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from django.utils import timezone
from django.conf import settings
import json
import logging
import os
import threading

from userpanel.models import WASenderSession, WASenderMessage, WASenderCampaign
from .wasender_service import WASenderService
from adminpanel.models import Subscription
from .moderation import evaluate_content

logger = logging.getLogger(__name__)


# ==================== Session Management Views ====================

@login_required
@never_cache  # Prevent browser from caching this page
def dashboard(request):
    """
    Main WA Sender dashboard
    Shows sessions, stats, and recent messages
    Supports multiple sessions per user with session switching
    OPTIMIZED: Async session status checks to prevent navigation lag
    REQUIRES: Active subscription to access
    ENFORCES: Content moderation restrictions
    """
    from whatsappapi.models import UserModerationProfile
    
    # Get moderation status for template context
    moderation_status = None
    try:
        mod_profile = UserModerationProfile.objects.get(user=request.user)
        
        # Permanent block - deny access completely
        if mod_profile.permanently_blocked:
            moderation_status = 'permanent'
            messages.error(
                request,
                "Your account has been permanently blocked. You cannot access the WhatsApp API dashboard."
            )
            return redirect('userpanel:dashboard')
        elif mod_profile.warnings_count >= 2:
            moderation_status = 'second'
        elif mod_profile.warnings_count >= 1:
            moderation_status = 'first'
    except UserModerationProfile.DoesNotExist:
        pass  # No moderation profile = clean user
    
    # Check if user has an active subscription
    subscription = Subscription.objects.filter(
        user=request.user,
        status='active',
        end_date__gt=timezone.now()
    ).first()
    
    subscription_is_active = subscription is not None
    
    # If no active subscription, redirect to pricing page
    if not subscription_is_active:
        messages.warning(request, "You need an active subscription to access the WhatsApp API Dashboard. Please purchase a plan to continue.")
        return redirect('userpanel:pricing')
    
    # Get all user's sessions (support multiple sessions)
    all_sessions = WASenderSession.objects.filter(user=request.user).order_by('-created_at')
    
    # Get active sessions (connected)
    active_sessions = all_sessions.filter(status='connected')
    
    # Check if user wants to switch to a specific session
    session_id_param = request.GET.get('session')
    if session_id_param:
        try:
            active_session = all_sessions.get(id=int(session_id_param))
        except (WASenderSession.DoesNotExist, ValueError):
            # Invalid session ID, fallback to most recent
            active_session = all_sessions.first()
    else:
        # Get the most recent session for display (could be any status)
        active_session = all_sessions.first()
    
    # OPTIMIZED: Only check session status on explicit refresh (not every page load)
    # Check if refresh parameter is present
    should_refresh = request.GET.get('refresh') == '1'
    
    if should_refresh and active_sessions.exists():
        service = WASenderService()
        for session in active_sessions[:3]:  # Check max 3 connected sessions
            try:
                # This will update status in DB if session is disconnected on WASender
                service.get_session_status(session)
                session.refresh_from_db()  # Reload from DB to get updated status
            except Exception as e:
                logger.error(f"Error updating session {session.session_id} status: {e}")
        
        # Refresh session lists after status check
        all_sessions = WASenderSession.objects.filter(user=request.user).order_by('-created_at')
        active_sessions = all_sessions.filter(status='connected')
        if session_id_param:
            try:
                active_session = all_sessions.get(id=int(session_id_param))
            except (WASenderSession.DoesNotExist, ValueError):
                active_session = all_sessions.first()
        else:
            active_session = all_sessions.first()
    
    # Calculate accurate stats from WASenderMessage model
    from datetime import datetime
    
    # Messages sent today (all statuses except failed)
    sent_today = WASenderMessage.objects.filter(
        user=request.user,
        created_at__date=timezone.now().date(),
        status__in=['sent', 'delivered', 'read', 'sending', 'queued']
    ).count()
    
    # Total messages ever sent (all statuses except failed)
    total_messages = WASenderMessage.objects.filter(
        user=request.user,
        status__in=['sent', 'delivered', 'read', 'sending', 'queued']
    ).count()
    
    # Messages sent this month (all statuses except failed)
    sent_this_month = WASenderMessage.objects.filter(
        user=request.user,
        created_at__month=timezone.now().month,
        created_at__year=timezone.now().year,
        status__in=['sent', 'delivered', 'read', 'sending', 'queued']
    ).count()

    # Moderation banner context for dashboard
    try:
        from .models import UserModerationProfile
        moderation_profile, _ = UserModerationProfile.objects.get_or_create(user=request.user)
    except Exception:
        moderation_profile = None

    context = {
        'active_session': active_session,
        'active_sessions': active_sessions,  # All connected sessions
        'all_sessions': all_sessions,
        'total_messages': total_messages,
        'sent_today': sent_today,
        'sent_this_month': sent_this_month,
        'moderation_profile': moderation_profile,
        'moderation_status': moderation_status,
        'subscription': subscription,
        'subscription_is_active': subscription_is_active,
    }
    
    return render(request, 'whatsappapi/dashboard.html', context)


@login_required
def sessions_list(request):
    """
    Show list of all WhatsApp sessions with management options
    Allows viewing, disconnecting, and deleting sessions
    OPTIMIZED: Only refresh status on explicit request to prevent blocking
    """
    # Get all sessions for the user (optimized with select_related to avoid N+1 queries)
    all_sessions = WASenderSession.objects.filter(user=request.user).select_related('user').order_by('-created_at')
    
    # OPTIMIZED: Only refresh status when explicitly requested (not every page load)
    should_refresh = request.GET.get('refresh') == '1'
    
    if should_refresh:
        connected_sessions = all_sessions.filter(status='connected')
        if connected_sessions.exists():
            service = WASenderService()
            for session in connected_sessions[:3]:  # Reduced to 3 for faster response
                try:
                    service.get_session_status(session)  # Updates status in DB
                except Exception as e:
                    logger.error(f"Error updating session {session.session_id} status: {e}")
            
            # Refresh session lists after status check
            all_sessions = WASenderSession.objects.filter(user=request.user).select_related('user').order_by('-created_at')
    
    # Categorize sessions
    active_sessions = all_sessions.filter(status='connected')
    pending_sessions = all_sessions.filter(status__in=['pending', 'connecting'])
    inactive_sessions = all_sessions.filter(status__in=['disconnected', 'error'])
    
    context = {
        'all_sessions': all_sessions,
        'active_sessions': active_sessions,
        'pending_sessions': pending_sessions,
        'inactive_sessions': inactive_sessions,
        'total_count': all_sessions.count(),
        'active_count': active_sessions.count(),
        'pending_count': pending_sessions.count(),
        'inactive_count': inactive_sessions.count(),
    }
    
    return render(request, 'whatsappapi/sessions.html', context)


@login_required
def create_session(request):
    """
    Create new WhatsApp session - Get phone number from user
    Supports multiple sessions per user
    REQUIRES: Active subscription
    """
    # Check if user has an active subscription
    subscription = Subscription.objects.filter(
        user=request.user,
        status='active',
        end_date__gt=timezone.now()
    ).first()
    
    if not subscription:
        messages.error(request, "You need an active subscription to create WhatsApp sessions. Please purchase a plan first.")
        return redirect('userpanel:pricing')
    
    if request.method == 'POST':
        phone_number = request.POST.get('phone_number', '').strip()
        
        if not phone_number:
            messages.error(request, "Phone number is required")
            return redirect('whatsappapi:dashboard')
        
        # Allow multiple sessions - just check if this phone number is already in use by THIS user
        existing_phone_session = WASenderSession.objects.filter(
            user=request.user,
            phone_number=phone_number,
            status__in=['connected', 'pending']
        ).first()
        
        if existing_phone_session:
            messages.warning(request, f"You already have a session with {phone_number}. Please use a different number or delete the existing session.")
            return redirect('whatsappapi:connect_session', session_id=existing_phone_session.id)
        
        # Clean up disconnected sessions for this phone number
        disconnected_sessions = WASenderSession.objects.filter(
            user=request.user,
            phone_number=phone_number,
            status='disconnected'
        )
        
        service = WASenderService()
        for old_session in disconnected_sessions:
            try:
                # Verify session still exists on WaSender API
                status_data = service.get_session_status(old_session)
                # If it exists on WaSender, keep it blocked
                if old_session.status in ['connected', 'pending']:
                    messages.warning(request, f"Session with {phone_number} is already active on WaSender. Please delete it first.")
                    return redirect('whatsappapi:connect_session', session_id=old_session.id)
            except Exception as e:
                logger.info(f"Disconnected session {old_session.session_id} no longer exists on WaSender, allowing cleanup")
                # Session doesn't exist on WaSender, safe to delete locally and create new one
                old_session.delete()
        
        try:
            # Auto-generate webhook URL for this user
            webhook_url = request.build_absolute_uri(
                reverse('whatsappapi:wasender_webhook', kwargs={'user_id': request.user.id})
            )
            
            # Log webhook URL generation
            logger.info(f"🔗 Auto-generated webhook URL for user {request.user.id}: {webhook_url}")
            
            # Create session with automatic webhook configuration
            session = service.create_session(request.user, webhook_url, phone_number)
            if session:
                # Initiate connection to get QR code
                service.connect_session(session)
                messages.success(request, f"Session created for {phone_number}! Please scan QR code to connect")
                return redirect('whatsappapi:connect_session', session_id=session.id)
            else:
                messages.error(request, "Failed to create session. Please try again or contact support.")
                return redirect('whatsappapi:dashboard')
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error creating session: {error_msg}")
            
            # Show user-friendly error messages
            if "Session limit reached" in error_msg or "upgrade your plan" in error_msg.lower():
                messages.error(
                    request, 
                    "You have reached your WhatsApp session limit on WASender. Please delete an existing session or upgrade your plan at https://wasenderapi.com/pricing"
                )
            elif "phone number has already been taken" in error_msg.lower():
                messages.error(
                    request, 
                    "This phone number is already registered with another session. Please delete the old session from WASender dashboard first."
                )
            else:
                messages.error(request, f"Failed to create session: {error_msg}")
            
            return redirect('whatsappapi:dashboard')
    
    return render(request, 'whatsappapi/dashboard.html')


@login_required
def connect_session(request, session_id):
    """
    Show QR code for WhatsApp connection
    """
    session = get_object_or_404(WASenderSession, id=session_id, user=request.user)
    
    service = WASenderService()
    
    # Always call connect first to ensure session is ready for QR
    if session.status in ['disconnected', 'logged_out']:
        logger.info(f"Session {session.session_id} is {session.status}, calling connect...")
        service.connect_session(session)
    
    # Get QR code if session is pending/need_scan
    qr_code = None
    if session.status in ['pending', 'need_scan']:
        try:
            qr_code = service.get_qr_code(session)
        except Exception as e:
            logger.error(f"Error getting QR code: {e}")
            messages.error(request, "Failed to get QR code. Please try again")
    
    context = {
        'session': session,
        'qr_code': qr_code,
    }
    
    return render(request, 'whatsappapi/connect.html', context)


@login_required
@require_POST
def check_session_status(request, session_id):
    """
    AJAX endpoint to check session connection status
    Polls WASender API to detect when QR code is scanned and session connects
    """
    session = get_object_or_404(WASenderSession, id=session_id, user=request.user)
    
    service = WASenderService()
    try:
        # Get latest status from WASender API
        status_data = service.get_session_status(session)
        session.refresh_from_db()  # Reload to get updated status
        
        # Build response with complete session data
        response_data = {
            'status': session.status,
            'phone_number': session.phone_number or '',
            'connected_phone_number': session.connected_phone_number or '',
            'connected': session.status == 'connected',
            'session_id': session.id,
            'data': status_data
        }
        
        # If session just connected, provide redirect URL with session parameter
        if session.status == 'connected':
            response_data['redirect_url'] = f"/whatsappapi/?session={session.id}"
        
        return JsonResponse(response_data)
    except Exception as e:
        logger.error(f"Error checking status for session {session_id}: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def refresh_all_sessions(request):
    """
    AJAX endpoint to refresh status of all user sessions
    Detects external disconnections from WASender API
    """
    service = WASenderService()
    all_sessions = WASenderSession.objects.filter(user=request.user).order_by('-created_at')[:10]
    
    updated_sessions = []
    for session in all_sessions:
        try:
            service.get_session_status(session)
            updated_sessions.append({
                'id': session.id,
                'session_id': session.session_id,
                'status': session.status,
                'phone_number': session.phone_number or '',
            })
        except Exception as e:
            logger.error(f"Error refreshing session {session.session_id}: {e}")
            updated_sessions.append({
                'id': session.id,
                'session_id': session.session_id,
                'status': session.status,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': True,
        'sessions': updated_sessions,
        'count': len(updated_sessions)
    })


@login_required
@require_POST
def disconnect_session(request, session_id):
    """
    Disconnect WhatsApp session
    """
    session = get_object_or_404(WASenderSession, id=session_id, user=request.user)
    
    service = WASenderService()
    try:
        if service.disconnect_session(session):
            messages.success(request, "Session disconnected successfully")
        else:
            # Check if it's a 502/503 error (API unavailable)
            messages.error(request, "Failed to disconnect. WaSender API may be temporarily unavailable. Please try again in a few moments.")
    except Exception as e:
        logger.error(f"Error disconnecting session: {e}")
        messages.error(request, "Error disconnecting session. Please try again.")
    
    return redirect('whatsappapi:dashboard')


@login_required
@require_POST
def delete_session(request, session_id):
    """
    Delete WhatsApp session
    """
    session = get_object_or_404(WASenderSession, id=session_id, user=request.user)
    session_id_text = session.session_id
    
    service = WASenderService()
    try:
        # Try to delete from WASender API
        # Returns True even if session doesn't exist (404)
        api_deleted = service.delete_session(session)
        
        # Always delete from local database
        session.delete()
        
        if api_deleted:
            messages.success(request, "Session deleted successfully")
        else:
            # API deletion failed but local deletion succeeded
            messages.warning(
                request,
                f"Session deleted from your account, but removal from WASender failed. "
                f"You may need to manually delete session {session_id_text} from WASender dashboard."
            )
    except Exception as e:
        logger.error(f"Error deleting session: {e}")
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('whatsappapi:dashboard')


# ==================== Message Sending Views ====================

def _get_send_campaign_context(user, session, **extra):
    """
    OPTIMIZED: Helper to build send_campaign context with optimized queries.
    Reduces duplicate code and prevents N+1 queries.
    """
    from .models import ContactList, CampaignTemplate
    
    context = {
        'session': session,
        'contact_lists': ContactList.objects.filter(user=user).select_related('user').only(
            'id', 'name', 'total_contacts', 'available_fields', 'created_at', 'user'
        ),
        'drafts': CampaignTemplate.objects.filter(
            user=user, status='draft'
        ).select_related('user', 'contact_list').only(
            'id', 'name', 'message_template', 'contact_list', 'created_at', 'user', 'status'
        ),
        'all_sessions': WASenderSession.objects.filter(user=user).select_related('user').only(
            'id', 'session_id', 'session_name', 'phone_number', 'connected_phone_number', 'status', 'user'
        ),
    }
    # Merge any extra context passed in
    context.update(extra)
    return context

@login_required
def send_campaign(request):
    """
    Send WhatsApp campaign view with draft templates and contact list support
    Supports multiple sessions - respects session selection from dashboard
    REQUIRES: Active subscription
    """
    from .models import ContactList, CampaignTemplate
    
    # Check if user has an active subscription
    subscription = Subscription.objects.filter(
        user=request.user,
        status='active',
        end_date__gt=timezone.now()
    ).first()
    
    if not subscription:
        messages.error(request, "You need an active subscription to send campaigns. Please purchase a plan first.")
        return redirect('userpanel:pricing')
    
    # Check if user selected a specific session from dashboard
    session_id_param = request.GET.get('session')
    
    if session_id_param:
        # User selected a specific session - use that one
        try:
            session = WASenderSession.objects.get(
                id=int(session_id_param),
                user=request.user,
                status='connected'
            )
        except (WASenderSession.DoesNotExist, ValueError):
            # Invalid or disconnected session, fallback to first connected
            session = WASenderSession.objects.filter(
                user=request.user,
                status='connected'
            ).first()
    else:
        # No specific session selected, use first connected session
        session = WASenderSession.objects.filter(
            user=request.user,
            status='connected'
        ).first()
    
    if not session:
        messages.error(request, "Please connect WhatsApp first")
        return redirect('whatsappapi:create_session')
    
    # Check for phone number mismatch (quick database-only check)
    if session.phone_number and session.connected_phone_number:
        if session.phone_number != session.connected_phone_number:
            messages.error(
                request, 
                f"Phone number mismatch! You registered with {session.phone_number} but scanned with {session.connected_phone_number}. "
                f"Please delete this session and create a new one with {session.connected_phone_number}."
            )
            # Do NOT redirect away; keep user on send page and show error (global toast)
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
    
    if request.method == 'POST':
        # CRITICAL: Get session from POST data (hidden field) to preserve user's selection
        session_id_post = request.POST.get('session_id')
        
        if session_id_post:
            try:
                # Use the session that was selected when form was loaded
                session = WASenderSession.objects.get(
                    id=int(session_id_post),
                    user=request.user,
                    status='connected'
                )
                logger.info(f"Using session {session.id} ({session.phone_number}) from form submission")
            except (WASenderSession.DoesNotExist, ValueError):
                messages.error(request, "Selected session is no longer available or disconnected")
                return redirect('whatsappapi:dashboard')
        else:
            # Fallback: no session in POST, use first connected (should not happen)
            session = WASenderSession.objects.filter(
                user=request.user,
                status='connected'
            ).first()
            if not session:
                messages.error(request, "No connected WhatsApp session found")
                return redirect('whatsappapi:dashboard')
        
        campaign_name = request.POST.get('campaign_name', '').strip()
        message = request.POST.get('message', '').strip()
        contact_list_id = request.POST.get('contact_list_id')
        attachment = request.FILES.get('attachment')
        compliance_confirmed = request.POST.get('compliance_confirmed') == 'true'

        # Admin enforcement: check user moderation profile (warnings / block)
        try:
            from .models import UserModerationProfile
            profile, _ = UserModerationProfile.objects.get_or_create(user=request.user)
            if profile.permanently_blocked:
                messages.error(request, "Your account has been permanently blocked by admin. You cannot send campaigns.")
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
            # Surface warning label to user (non-blocking)
            if profile.warnings_count:
                messages.warning(
                    request,
                    f"Admin Warning: {getattr(profile, 'warning_label', f'Level {profile.warnings_count}')} active on your account. Please ensure campaigns strictly follow policy."
                )
        except Exception:
            # Fail-open if profile cannot be loaded
            pass

        # Enforce compliance confirmation attestation
        if not compliance_confirmed:
            messages.error(request, "Please confirm compliance with WhatsApp policy and sending rules before sending.")
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))

        # Normalize monospace formatting to WhatsApp-compatible backticks
        try:
            import re
            message = re.sub(r"'''([\s\S]*?)'''", r"```\1```", message)
            message = message.replace('‘', "'").replace('’', "'").replace('´', "'")
            message = re.sub(r"[’‘]{3}([\s\S]*?)[’‘]{3}", r"```\1```", message)
        except Exception:
            pass
        
        # Validate campaign name format (only lowercase, numbers, underscores)
        import re
        if not re.match(r'^[a-z0-9_]+$', campaign_name):
            messages.error(request, "Campaign name must contain only lowercase letters, numbers, and underscores (e.g., new_year_offer_2025)")
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
        
        if not campaign_name or not message or not contact_list_id:
            messages.error(request, "Campaign name, message, and contact list are required")
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
        
        # Get contact list
        contact_list = ContactList.objects.filter(id=contact_list_id, user=request.user).first()
        if not contact_list:
            messages.error(request, "Invalid contact list selected")
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
        
        # Get all contacts from the list
        from .models import Contact
        contacts = Contact.objects.filter(contact_list=contact_list)
        
        if not contacts.exists():
            messages.error(request, "Contact list is empty")
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))

        # Content moderation gate: evaluate message before uploading attachments or queuing
        moderation = evaluate_content(message, attachment_type=None)
        # AI-only gate: treat review as block when enabled
        _ai_only_val = os.environ.get('AI_ONLY_GATE', None)
        if _ai_only_val is None:
            try:
                # Default to true: treat review as block in the UI gate
                _ai_only_val = getattr(settings, 'AI_ONLY_GATE', 'true')
            except Exception:
                _ai_only_val = 'true'
        ai_only_gate = str(_ai_only_val).lower() in {'1', 'true', 'yes'}
        if moderation.get('blocked') or (ai_only_gate and moderation.get('requires_review')):
            # Format human-readable reasons for the Django message banner
            try:
                _reasons = moderation.get('reasons') or []
                _formatted = []
                for r in _reasons:
                    label, _, items = (r or '').partition(':')
                    label = (label or '').replace('_', ' ').strip()
                    if items:
                        _formatted.append(f"{label} ({items})")
                    elif label:
                        _formatted.append(label)
                reason_summary = "; ".join(_formatted[:6])
                risk = int(moderation.get('risk_score') or 0)
                msg_text = (
                    f"Blocked by content moderation. Reasons: {reason_summary}. "
                    f"Risk score: {risk}. Please revise to comply with WhatsApp and local laws."
                )
            except Exception:
                msg_text = (
                    'Your message appears to contain illegal or prohibited content and was blocked. '
                    'Please revise to comply with WhatsApp and local laws.'
                )
            # Log incident for admin review and action
            try:
                from .models import ModerationIncident
                ModerationIncident.objects.create(
                    user=request.user,
                    campaign_name=campaign_name or '(unnamed)'
                    ,contact_list=contact_list,
                    message_text=message,
                    status='blocked',
                    risk_score=int(moderation.get('risk_score') or 0),
                    reasons_text='; '.join(moderation.get('reasons') or [])[:2000]
                )
                logger.info(f"✅ ModerationIncident created successfully for user {request.user.email}")
            except Exception as e:
                logger.error(f"❌ FAILED to create ModerationIncident: {type(e).__name__}: {str(e)}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
            messages.error(request, msg_text)
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(
                request.user, session, moderation=moderation
            ))
        elif moderation.get('requires_review'):
            try:
                _reasons = moderation.get('reasons') or []
                _formatted = []
                for r in _reasons:
                    label, _, items = (r or '').partition(':')
                    label = (label or '').replace('_', ' ').strip()
                    if items:
                        _formatted.append(f"{label} ({items})")
                    elif label:
                        _formatted.append(label)
                reason_summary = "; ".join(_formatted[:6])
                risk = int(moderation.get('risk_score') or 0)
                warn_text = (
                    f"Content flagged for review. Reasons: {reason_summary}. Risk score: {risk}. "
                    f"Please double-check before sending."
                )
            except Exception:
                warn_text = 'Your message may violate content policies and was flagged; please double-check before sending.'
            # Log review incident for admin visibility
            try:
                from .models import ModerationIncident
                ModerationIncident.objects.create(
                    user=request.user,
                    campaign_name=campaign_name or '(unnamed)'
                    ,contact_list=contact_list,
                    message_text=message,
                    status='review',
                    risk_score=int(moderation.get('risk_score') or 0),
                    reasons_text='; '.join(moderation.get('reasons') or [])[:2000]
                )
                logger.info(f"✅ ModerationIncident (review) created successfully for user {request.user.email}")
            except Exception as e:
                logger.error(f"❌ FAILED to create ModerationIncident (review): {type(e).__name__}: {str(e)}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
            messages.warning(request, warn_text)
            # Continue flow but keep moderation info in context for UI hinting
            
        
        # Handle attachment - store file temporarily, upload in background
        import uuid
        
        attachment_file_path = None
        attachment_filename = None
        attachment_type = None
        
        if attachment:
            # Determine attachment type by file extension
            file_ext = attachment.name.lower().split('.')[-1]
            if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
                attachment_type = 'image'
            elif file_ext in ['mp4', 'avi', 'mov', '3gp']:
                attachment_type = 'video'
            elif file_ext in ['mp3', 'wav', 'ogg', 'aac', 'm4a']:
                attachment_type = 'audio'
            else:
                attachment_type = 'document'
            
            # Save to temporary location for background upload
            # Create temp directory if not exists
            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_uploads')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Generate unique filename
            temp_filename = f"{uuid.uuid4()}_{attachment.name}"
            attachment_file_path = os.path.join(temp_dir, temp_filename)
            attachment_filename = attachment.name
            
            # Save file to temp location
            try:
                with open(attachment_file_path, 'wb+') as destination:
                    for chunk in attachment.chunks():
                        destination.write(chunk)
                logger.info(f"Attachment saved to temp: {attachment_file_path}")
            except Exception as e:
                logger.error(f"Failed to save temp attachment: {e}")
                messages.error(request, f"Failed to save attachment: {str(e)}")
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
        
        # Check for duplicate campaign submission (same name, session, and contact list within last 5 minutes)
        from datetime import timedelta
        
        # CRITICAL: Single session mode - only ONE campaign across ALL sessions
        # This prevents WhatsApp from flagging simultaneous sends as spam and logging out sessions
        any_active_campaign = WASenderCampaign.objects.filter(
            user=request.user,
            status__in=['pending', 'running']
        ).exists()
        
        if any_active_campaign:
            # Show warning on the send page instead of redirecting
            from userpanel.models import WASenderCampaign as CampaignModel
            active_campaigns = CampaignModel.objects.filter(
                user=request.user,
                status__in=['pending', 'running']
            ).order_by('-created_at')
            
            active_campaign = active_campaigns.first()
            session_name = active_campaign.session.session_name if active_campaign and active_campaign.session else "Unknown"
            
            messages.warning(
                request, 
                f"A campaign is already running on session '{session_name}'. Please wait until it completes before starting a new campaign. This prevents WhatsApp from logging out your sessions."
            )
            return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(
                request.user, session, 
                session_has_active_campaign=True,
                active_campaigns=active_campaigns
            ))
        
        # Create campaign with pending status
        # Get advanced controls from POST if enabled
        use_advanced_controls = request.POST.get('use_advanced_controls') == 'true'
        
        # Extract user-defined values or use defaults from settings.py
        if use_advanced_controls:
            # User enabled advanced controls - use their custom values
            random_delay_min = int(request.POST.get('random_delay_min', settings.DEFAULT_RANDOM_DELAY_MIN))
            random_delay_max = int(request.POST.get('random_delay_max', settings.DEFAULT_RANDOM_DELAY_MAX))
            batch_size_min = int(request.POST.get('batch_size_min', settings.DEFAULT_BATCH_SIZE_MIN))
            batch_size_max = int(request.POST.get('batch_size_max', settings.DEFAULT_BATCH_SIZE_MAX))
            batch_cooldown_min = float(request.POST.get('batch_cooldown_min', settings.DEFAULT_BATCH_COOLDOWN_MIN))
            batch_cooldown_max = float(request.POST.get('batch_cooldown_max', settings.DEFAULT_BATCH_COOLDOWN_MAX))
            
            # Validate Advanced Controls input
            if random_delay_min < 3:
                messages.error(request, "❌ Minimum delay must be at least 3 seconds for account safety")
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
            
            if random_delay_max < random_delay_min:
                messages.error(request, "❌ Maximum delay must be greater than minimum delay")
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
            
            if random_delay_max > 60:
                messages.error(request, "❌ Maximum delay cannot exceed 60 seconds (too slow)")
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
            
            # Validate batch sizes (if batching is enabled)
            if batch_size_max > 0:
                if batch_size_min < 10:
                    messages.error(request, "❌ Minimum batch size must be at least 10 contacts")
                    return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
                
                if batch_size_max < batch_size_min:
                    messages.error(request, "❌ Maximum batch size must be greater than minimum batch size")
                    return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
                
                if batch_size_max > 500:
                    messages.error(request, "❌ Maximum batch size cannot exceed 500 contacts")
                    return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
            
            # Validate cooldowns (if cooldown is enabled)
            if batch_cooldown_max > 0:
                if batch_cooldown_min < 1:
                    messages.error(request, "❌ Minimum cooldown must be at least 1 minute")
                    return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
                
                if batch_cooldown_max < batch_cooldown_min:
                    messages.error(request, "❌ Maximum cooldown must be greater than minimum cooldown")
                    return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
                
                if batch_cooldown_max > 120:
                    messages.error(request, "❌ Maximum cooldown cannot exceed 120 minutes (2 hours)")
                    return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
            
            logger.info(f"✅ Advanced Controls ENABLED for campaign '{campaign_name}'")
            logger.info(f"   Delay Range: {random_delay_min}-{random_delay_max} seconds")
            logger.info(f"   Batch Size Range: {batch_size_min}-{batch_size_max} contacts")
            logger.info(f"   Cooldown Range: {batch_cooldown_min}-{batch_cooldown_max} minutes")
            logger.info(f"✅ Advanced Controls validation passed")
        else:
            # Advanced controls disabled - use standard mode (NO batching, fixed delays)
            # Standard mode uses MESSAGE_DELAY_WITH_PROTECTION or MESSAGE_DELAY_WITHOUT_PROTECTION
            # based on session's account_protection_enabled setting (handled in tasks.py)
            random_delay_min = 0  # Not used in standard mode
            random_delay_max = 0  # Not used in standard mode
            batch_size_min = 0    # Disable batching in standard mode
            batch_size_max = 0    # Disable batching in standard mode
            batch_cooldown_min = 0  # No cooldown in standard mode
            batch_cooldown_max = 0  # No cooldown in standard mode
            
            logger.info(f"ℹ️ STANDARD MODE for campaign '{campaign_name}'")
            logger.info(f"   Will use fixed delays based on session protection setting:")
            logger.info(f"   - Protection ON: {settings.MESSAGE_DELAY_WITH_PROTECTION}s")
            logger.info(f"   - Protection OFF: {settings.MESSAGE_DELAY_WITHOUT_PROTECTION}s")
            logger.info(f"   No batching, no cooldowns")
        
        campaign = WASenderCampaign.objects.create(
            user=request.user,
            session=session,
            name=campaign_name,
            message_template=message,
            contact_list=contact_list,
            recipients=[],
            total_recipients=contacts.count(),
            status='pending',
            attachment_url=attachment_file_path,  # Store temp path temporarily
            attachment_type=attachment_type,
            attachment_public_id=attachment_filename,  # Store original filename temporarily
            wasender_document_url=None,
            # Advanced sending controls
            use_advanced_controls=use_advanced_controls,
            random_delay_min=random_delay_min,
            random_delay_max=random_delay_max,
            batch_size_min=batch_size_min,
            batch_size_max=batch_size_max,
            batch_cooldown_min=batch_cooldown_min,
            batch_cooldown_max=batch_cooldown_max
        )
        
        # Queue background task to send campaign (idempotent at task start)
        try:
            from django_q.tasks import async_task
            # Pass callable directly to avoid import resolution issues on some Windows paths
            from whatsappapi.tasks import send_campaign_async
            task_id = async_task(
                send_campaign_async,
                campaign.id,
                task_name=f"campaign_{campaign.id}_{campaign.name}"
            )
        except Exception as e:
            # Primary queueing failed — attempt local thread fallback to keep UX smooth
            logger.error(f"Failed to queue background task for campaign {campaign.id}: {e}")
            try:
                from whatsappapi.tasks import send_campaign_async as _runner
                thread_name = f"campaign_{campaign.id}_worker"
                worker = threading.Thread(target=_runner, args=(campaign.id,), name=thread_name, daemon=True)
                worker.start()
                task_id = f"thread:{thread_name}"
                campaign.task_id = task_id
                campaign.save(update_fields=['task_id'])
                logger.warning(
                    f"Django-Q queue failed; started local thread '{thread_name}' for campaign {campaign.id}."
                )
                messages.success(
                    request,
                    f"Campaign '{campaign_name}' has been started in the background (local fallback)."
                )
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(
                    request.user, session,
                    queue_success_modal=True,
                    queue_success_message=f"Campaign '{campaign_name}' has been started in the background (local fallback).",
                    queued_campaign=campaign
                ))
            except Exception as e2:
                logger.error(
                    f"Fallback thread start failed for campaign {campaign.id}: {e2}"
                )
                messages.error(
                    request,
                    (
                        "Failed to queue background task and local fallback. "
                        f"Please try again or contact support. Error: {str(e2)}"
                    )
                )
                # Mark campaign as failed to avoid leaving it stuck in pending
                try:
                    campaign.status = 'failed'
                    campaign.save(update_fields=['status'])
                except Exception:
                    pass
                return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(request.user, session))
        
        # Save task ID to campaign
        campaign.task_id = task_id
        campaign.save()
        
        logger.info(f"Campaign '{campaign_name}' queued for background processing (Task ID: {task_id})")
        # Render success modal on send page instead of redirect
        messages.success(request, f"Campaign '{campaign_name}' has been queued! It will be sent in the background. Check the campaigns page for progress.")
        return render(request, 'whatsappapi/send_campaign.html', _get_send_campaign_context(
            request.user, session,
            queue_success_modal=True,
            queue_success_message=f"Campaign '{campaign_name}' has been queued! It will be sent in the background.",
            queued_campaign=campaign
        ))
    
    # GET request - show form
    draft_id = request.GET.get('draft')
    selected_list_id = request.GET.get('list')
    selected_draft = None
    selected_contact_list = None
    
    # Resolve preselected contact list from query param first (overrides draft)
    if draft_id:
        selected_draft = CampaignTemplate.objects.filter(
            id=draft_id, user=request.user
        ).select_related('user', 'contact_list').first()

    if selected_list_id:
        selected_contact_list = ContactList.objects.filter(
            id=selected_list_id, user=request.user
        ).select_related('user').first()

    # Get available fields from selected contact list or draft's contact list
    available_fields = []
    if selected_contact_list:
        available_fields = selected_contact_list.available_fields if selected_contact_list.available_fields else []
    elif selected_draft and selected_draft.contact_list:
        available_fields = selected_draft.contact_list.available_fields if selected_draft.contact_list.available_fields else []
    
    # Use the helper function to get optimized context
    context = _get_send_campaign_context(
        request.user, session,
        selected_draft=selected_draft,
        available_fields=available_fields,
        selected_contact_list=selected_contact_list,
        selected_contact_list_id=selected_contact_list.id if selected_contact_list else None
    )
    
    return render(request, 'whatsappapi/send_campaign.html', context)


@login_required
@require_POST
def ai_draft(request):
    """
    Generate AI-powered campaign message draft
    Uses the ai_draft module for cleaner separation of concerns
    """
    try:
        # Import AI draft service
        from .ai_draft import generate_ai_draft, generate_marketing_guide
        
        # Parse request body
        try:
            if request.content_type and 'application/json' in request.content_type:
                payload = json.loads(request.body or '{}')
            else:
                payload = request.POST
        except Exception:
            payload = request.POST

        # Extract parameters
        prompt = (payload.get('prompt') or '').strip()
        variables = payload.get('variables') or []
        campaign_name = (payload.get('campaign_name') or '').strip()
        tone = ((payload.get('tone') or '').strip().lower())
        length = ((payload.get('length') or '').strip().lower())

        # Normalize variables to list
        if isinstance(variables, str):
            try:
                variables = json.loads(variables)
            except Exception:
                variables = [variables]
        variables = [str(v).strip() for v in (variables or []) if str(v).strip()]

        # Moderate the prompt before sending to AI
        prompt_mod = evaluate_content(prompt)
        if prompt_mod.get('blocked'):
            _reasons = prompt_mod.get('reasons') or []
            _formatted = []
            for r in _reasons:
                label, _, items = (r or '').partition(':')
                label = (label or '').replace('_', ' ').strip()
                if items:
                    _formatted.append(f"{label} ({items})")
                elif label:
                    _formatted.append(label)
            reason_text = '; '.join(_formatted[:6])
            return JsonResponse({
                'error': f'Prompt blocked by content moderation. Reasons: {reason_text}',
                'reasons': _reasons,
                'risk_score': int(prompt_mod.get('risk_score') or 0)
            }, status=403)

        # Generate AI draft using the dedicated module
        result = generate_ai_draft(
            prompt=prompt,
            tone=tone,
            length=length,
            variables=variables,
            campaign_name=campaign_name
        )
        
        # Check if generation was successful
        if not result.get('success'):
            return JsonResponse({'error': result.get('error', 'Unknown error')}, status=500)
        
        cleaned_draft = result.get('draft', '')
        
        # Moderate the generated draft
        gen_mod = evaluate_content(cleaned_draft)
        if gen_mod.get('blocked'):
            _reasons = gen_mod.get('reasons') or []
            _formatted = []
            for r in _reasons:
                label, _, items = (r or '').partition(':')
                label = (label or '').replace('_', ' ').strip()
                if items:
                    _formatted.append(f"{label} ({items})")
                elif label:
                    _formatted.append(label)
            reason_text = '; '.join(_formatted[:6])
            return JsonResponse({
                'error': f'Draft blocked by moderation. Reasons: {reason_text}',
                'reasons': _reasons,
                'risk_score': int(gen_mod.get('risk_score') or 0)
            }, status=403)

        return JsonResponse({
            'success': True,
            'draft': cleaned_draft,
            'marketing_guide': generate_marketing_guide(cleaned_draft, tone, campaign_name),
            'moderation': {
                'requires_review': gen_mod.get('requires_review', False),
                'risk_score': int(gen_mod.get('risk_score') or 0)
            }
        })
    except Exception as e:
        logger.error(f"AI draft error: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_POST
def send_single_message(request):
    """
    AJAX endpoint to send single message
    """
    session = WASenderSession.objects.filter(
        user=request.user,
        status='connected'
    ).first()
    
    if not session:
        return JsonResponse({'error': 'No active session'}, status=400)
    
    data = json.loads(request.body)
    recipient = data.get('recipient', '').strip()
    message = data.get('message', '').strip()
    
    if not recipient or not message:
        return JsonResponse({'error': 'Recipient and message required'}, status=400)
    
    # Content moderation for single message
    mod = evaluate_content(message)
    # AI-only gate: treat review as block when enabled
    _ai_only_val = os.environ.get('AI_ONLY_GATE', None)
    if _ai_only_val is None:
        try:
            _ai_only_val = getattr(settings, 'AI_ONLY_GATE', '')
        except Exception:
            _ai_only_val = ''
    ai_only_gate = str(_ai_only_val).lower() in {'1', 'true', 'yes'}
    if mod.get('blocked') or (ai_only_gate and mod.get('requires_review')):
        # Provide human-readable reasons and risk score in JSON for frontend display
        _reasons = mod.get('reasons') or []
        _formatted = []
        for r in _reasons:
            label, _, items = (r or '').partition(':')
            label = (label or '').replace('_', ' ').strip()
            if items:
                _formatted.append(f"{label} ({items})")
            elif label:
                _formatted.append(label)
        reason_text = "; ".join(_formatted[:6])
        return JsonResponse(
            {
                'error': f'Message blocked by content moderation. Reasons: {reason_text}',
                'reasons': _reasons,
                'risk_score': int(mod.get('risk_score') or 0),
                'decision': int(mod.get('decision', 0)),
                'policy_hits': mod.get('policy_hits') or []
            },
            status=403
        )
    
    
    service = WASenderService()
    try:
        msg = service.send_text_message(session, recipient, message)
        if msg:
            try:
                current_meta = msg.metadata or {}
                current_meta.update({'moderation_flagged': mod.get('requires_review', False), 'risk_score': mod.get('risk_score', 0)})
                msg.metadata = current_meta
                msg.save(update_fields=['metadata'])
            except Exception:
                pass
            return JsonResponse({
                'success': True,
                'message_id': msg.message_id,
                'status': msg.status,
                'moderation': {
                    'risk_score': int(mod.get('risk_score') or 0),
                    'requires_review': bool(mod.get('requires_review')), 
                    'decision': int(mod.get('decision', 1)),
                    'reasons': mod.get('reasons') or [],
                    'policy_hits': mod.get('policy_hits') or []
                }
            })
        else:
            return JsonResponse({'error': 'Failed to send'}, status=500)
    except Exception as e:
        logger.error(f"Error sending message: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def send_test_message(request):
    """
    AJAX endpoint to send test campaign message
    Sends message without saving campaign
    """
    session = WASenderSession.objects.filter(
        user=request.user,
        status='connected'
    ).first()
    
    if not session:
        return JsonResponse({'error': 'No active session'}, status=400)
    
    try:
        # Get test parameters
        test_phone = request.POST.get('test_phone', '').strip()
        test_message = request.POST.get('test_message', '').strip()
        test_attachment = request.FILES.get('test_attachment')
        
        if not test_phone:
            return JsonResponse({'error': 'Phone number required'}, status=400)
        
        if not test_message:
            return JsonResponse({'error': 'Message required'}, status=400)
        
        # Content moderation for test message
        mod = evaluate_content(test_message)
        # AI-only gate: treat review as block when enabled
        _ai_only_val = os.environ.get('AI_ONLY_GATE', None)
        if _ai_only_val is None:
            try:
                _ai_only_val = getattr(settings, 'AI_ONLY_GATE', '')
            except Exception:
                _ai_only_val = ''
        ai_only_gate = str(_ai_only_val).lower() in {'1', 'true', 'yes'}
        if mod.get('blocked') or (ai_only_gate and mod.get('requires_review')):
            _reasons = mod.get('reasons') or []
            _formatted = []
            for r in _reasons:
                label, _, items = (r or '').partition(':')
                label = (label or '').replace('_', ' ').strip()
                if items:
                    _formatted.append(f"{label} ({items})")
                elif label:
                    _formatted.append(label)
            reason_text = "; ".join(_formatted[:6])
            return JsonResponse(
                {
                    'error': f'Test message blocked by content moderation. Reasons: {reason_text}',
                    'reasons': _reasons,
                    'risk_score': int(mod.get('risk_score') or 0)
                },
                status=403
            )
        
        service = WASenderService()
        
        # If attachment provided, upload to Cloudinary first
        if test_attachment:
            import cloudinary.uploader
            
            # Determine attachment type
            file_ext = test_attachment.name.lower().split('.')[-1]
            if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
                attachment_type = 'image'
                resource_type = 'image'
            elif file_ext in ['mp4', 'avi', 'mov', '3gp']:
                attachment_type = 'video'
                resource_type = 'video'
            elif file_ext in ['mp3', 'wav', 'ogg', 'aac', 'm4a']:
                attachment_type = 'audio'
                resource_type = 'video'
            else:
                attachment_type = 'document'
                resource_type = 'raw'
            
            # Get clean filename without extension for public_id
            clean_filename = os.path.splitext(test_attachment.name)[0]
            # Remove spaces and special characters
            clean_filename = clean_filename.replace(' ', '_')
            
            # Upload to Cloudinary with original filename
            upload_result = cloudinary.uploader.upload(
                test_attachment,
                folder=f"wa_test/{request.user.id}",
                resource_type=resource_type,
                type='upload',  # force public delivery
                access_mode='public',
                public_id=clean_filename,  # Use original filename
                use_filename=False,  # Don't use default filename
                unique_filename=False,
                overwrite=True
            )
            
            attachment_url = upload_result['secure_url']
            logger.info(f"Test attachment uploaded to Cloudinary: {attachment_url}")
            
            # Send media message
            msg = service.send_media_message(
                session=session,
                recipient=test_phone,
                media_url=attachment_url,
                message_type=attachment_type,
                caption=test_message,
                public_id=upload_result.get('public_id')
            )
        else:
            # Send text message
            msg = service.send_text_message(session, test_phone, test_message)

        if msg and msg.status == 'sent':
            try:
                current_meta = msg.metadata or {}
                current_meta.update({'moderation_flagged': mod.get('requires_review', False), 'risk_score': mod.get('risk_score', 0)})
                msg.metadata = current_meta
                msg.save(update_fields=['metadata'])
            except Exception:
                pass
            return JsonResponse({
                'success': True,
                'message': f'Test message sent to {test_phone}',
                'message_id': msg.message_id
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Failed to send test message'
            }, status=500)
            
    except Exception as e:
        logger.error(f"Error sending test message: {e}")
        return JsonResponse({'error': str(e)}, status=500)


# ==================== Webhook Handler ====================

@csrf_exempt
@require_POST
def wasender_webhook(request, user_id):
    """
    Webhook endpoint for WASender callbacks
    Receives message status updates and connection events
    
    Webhook URL format: https://yourdomain.com/whatsappapi/webhook/{user_id}/
    
    Events handled:
    - messages.received: Incoming messages
    - messages.update: Message status updates (sent, delivered, read, failed)
    - session.status: Session connection status changes
    """
    import time
    start_time = time.time()
    
    try:
        # Parse payload
        payload = json.loads(request.body)
        event_type = payload.get('event', 'unknown')
        session_id = payload.get('session_id', 'unknown')
        
        # Log webhook receipt with full details
        logger.info(f"📥 WEBHOOK RECEIVED | User: {user_id} | Event: {event_type} | Session: {session_id}")
        logger.info(f"📦 Webhook Payload: {json.dumps(payload, indent=2)}")
        
        # Process webhook
        service = WASenderService()
        result = service.process_webhook(payload)
        
        # Calculate processing time
        processing_time = (time.time() - start_time) * 1000  # Convert to ms
        
        if result:
            logger.info(f"✅ WEBHOOK PROCESSED | User: {user_id} | Event: {event_type} | Time: {processing_time:.2f}ms")
        else:
            logger.warning(f"⚠️ WEBHOOK PROCESSING FAILED | User: {user_id} | Event: {event_type}")
        
        return JsonResponse({
            'status': 'ok',
            'event': event_type,
            'processed': result,
            'processing_time_ms': round(processing_time, 2)
        })
        
    except json.JSONDecodeError as e:
        logger.error(f"❌ WEBHOOK JSON ERROR | User: {user_id} | Error: {e} | Body: {request.body}")
        return JsonResponse({'error': 'Invalid JSON payload'}, status=400)
    except Exception as e:
        logger.error(f"❌ WEBHOOK CRITICAL ERROR | User: {user_id} | Error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


# ==================== Campaign Management Views ====================

@login_required
def campaign_list(request):
    """
    List all campaigns with pagination (12 per page)
    Dynamically calculate stats from messages for real-time updates
    """
    campaigns_list = WASenderCampaign.objects.filter(user=request.user).order_by('-created_at')
    
    # Add pagination - 12 items per page
    from django.core.paginator import Paginator
    paginator = Paginator(campaigns_list, 12)  # Show 12 campaigns per page
    
    page_number = request.GET.get('page')
    campaigns = paginator.get_page(page_number)
    
    # Calculate real-time stats from messages for each campaign
    from django.db.models import Q, Count
    for campaign in campaigns:
        # Get messages by metadata campaign_id first (newer method)
        messages_qs = WASenderMessage.objects.filter(
            metadata__campaign_id=campaign.id
        )
        
        # Fallback to time-based if no metadata
        if not messages_qs.exists():
            messages_qs = WASenderMessage.objects.filter(
                session=campaign.session,
                created_at__gte=campaign.created_at
            )
            # If campaign has recipients list, filter by them
            if campaign.recipients:
                phones = [r.get('phone') for r in campaign.recipients if r.get('phone')]
                if phones:
                    messages_qs = messages_qs.filter(recipient__in=phones)
        
        # Calculate counts from actual messages
        sent_count = messages_qs.filter(status__in=['sent', 'delivered', 'read']).count()
        failed_count = messages_qs.filter(status='failed').count()
        delivered_count = messages_qs.filter(status__in=['delivered', 'read']).count()
        
        # Update campaign object with current stats (for display)
        campaign.messages_sent = sent_count
        campaign.messages_failed = failed_count
        campaign.messages_delivered = delivered_count
        
        # Also save to database if stats changed (for persistence)
        if campaign.messages_sent != sent_count or campaign.messages_failed != failed_count:
            try:
                WASenderCampaign.objects.filter(id=campaign.id).update(
                    messages_sent=sent_count,
                    messages_failed=failed_count,
                    messages_delivered=delivered_count
                )
            except Exception:
                pass
    
    context = {
        'campaigns': campaigns,
    }
    
    return render(request, 'whatsappapi/campaigns.html', context)


@login_required
def campaign_detail(request, campaign_id):
    """
    Campaign detail and statistics
    Dynamically calculate stats from messages for real-time updates
    """
    from django.core.paginator import Paginator
    
    campaign = get_object_or_404(WASenderCampaign, id=campaign_id, user=request.user)
    
    # Calculate real-time stats from messages
    # Try metadata-based query first (most accurate for new campaigns)
    messages_qs_metadata = WASenderMessage.objects.filter(
        metadata__campaign_id=campaign.id
    )
    
    # Always also get time-based query (fallback for old campaigns)
    messages_qs_time = WASenderMessage.objects.filter(
        session=campaign.session,
        created_at__gte=campaign.created_at
    )
    
    # Use metadata query if it has messages, otherwise use time-based
    # This ensures we always get stats even for old campaigns
    messages_qs = messages_qs_metadata if messages_qs_metadata.exists() else messages_qs_time
    
    # Update campaign object with real-time stats
    campaign.messages_sent = messages_qs.filter(status__in=['sent', 'delivered', 'read']).count()
    campaign.messages_delivered = messages_qs.filter(status__in=['delivered', 'read']).count()
    campaign.messages_read = messages_qs.filter(status='read').count()
    campaign.messages_failed = messages_qs.filter(status='failed').count()
    
    # Also save to database for persistence
    try:
        WASenderCampaign.objects.filter(id=campaign.id).update(
            messages_sent=campaign.messages_sent,
            messages_failed=campaign.messages_failed,
            messages_delivered=campaign.messages_delivered
        )
    except Exception:
        pass
    
    # Base queryset: messages for this session
    messages_queryset = WASenderMessage.objects.filter(
        session=campaign.session
    ).order_by('-created_at')
    
    # Prefer metadata linkage if available (newer messages tagged with campaign_id)
    # Use JSON key transform for compatibility with SQLite (avoid __contains)
    linked_qs = messages_queryset.filter(metadata__campaign_id=campaign.id)
    if linked_qs.exists():
        messages_queryset = linked_qs
    else:
        # Fallback to time window starting at campaign.created_at
        messages_queryset = messages_queryset.filter(created_at__gte=campaign.created_at)
    
    # Further filter by campaign recipients if available and not using metadata-linkage
    if not linked_qs.exists():
        if campaign.recipients:
            phone_numbers = [recipient.get('phone') for recipient in campaign.recipients if recipient.get('phone')]
            if phone_numbers:
                unique_phone_numbers = list(set(phone_numbers))
                messages_queryset = messages_queryset.filter(recipient__in=unique_phone_numbers)
        else:
            # SQLite does not support DISTINCT ON; emulate latest-per-recipient via Subquery
            try:
                from django.db.models import OuterRef, Subquery
                latest_ids = WASenderMessage.objects.filter(
                    session=campaign.session,
                    recipient=OuterRef('recipient'),
                    created_at__gte=campaign.created_at
                ).order_by('-created_at').values('id')[:1]
                messages_queryset = messages_queryset.filter(id=Subquery(latest_ids)).order_by('-created_at')
            except Exception:
                # As a safe fallback, just order by newest
                messages_queryset = messages_queryset.order_by('-created_at')
    
    # Paginate messages (10 per page)
    paginator = Paginator(messages_queryset, 10)
    page_number = request.GET.get('page')
    messages_page = paginator.get_page(page_number)
    
    context = {
        'campaign': campaign,
        'campaign_messages': messages_page,
    }
    
    return render(request, 'whatsappapi/campaign_detail.html', context)


@login_required
def campaign_stats_api(request, campaign_id):
    """
    API endpoint to get real-time campaign statistics without full page reload
    Returns JSON with messages_sent, messages_failed, total_recipients, and status
    """
    from django.http import JsonResponse
    
    campaign = get_object_or_404(WASenderCampaign, id=campaign_id, user=request.user)
    
    # Calculate real-time stats from messages
    # Try metadata-based query first (most accurate for new campaigns)
    messages_qs_metadata = WASenderMessage.objects.filter(
        metadata__campaign_id=campaign.id
    )
    
    # Always also get time-based query (fallback for old campaigns)
    messages_qs_time = WASenderMessage.objects.filter(
        session=campaign.session,
        created_at__gte=campaign.created_at
    )
    
    # Use metadata query if it has messages, otherwise use time-based
    messages_qs = messages_qs_metadata if messages_qs_metadata.exists() else messages_qs_time
    
    # Calculate counts by status
    messages_sent = messages_qs.filter(status__in=['sent', 'delivered', 'read']).count()
    messages_delivered = messages_qs.filter(status__in=['delivered', 'read']).count()
    messages_read = messages_qs.filter(status='read').count()
    messages_failed = messages_qs.filter(status='failed').count()
    
    # Calculate success rate
    success_rate = 0
    if campaign.total_recipients > 0:
        success_rate = round((messages_delivered / campaign.total_recipients) * 100, 2)
    
    # Update database with latest stats
    try:
        WASenderCampaign.objects.filter(id=campaign.id).update(
            messages_sent=messages_sent,
            messages_delivered=messages_delivered,
            messages_read=messages_read,
            messages_failed=messages_failed
        )
    except Exception:
        pass
    
    # Refresh campaign data from DB to get latest cooldown info
    campaign.refresh_from_db()
    
    return JsonResponse({
        'messages_sent': messages_sent,
        'messages_delivered': messages_delivered,
        'messages_read': messages_read,
        'messages_failed': messages_failed,
        'success_rate': success_rate,
        'total_recipients': campaign.total_recipients,
        'status': campaign.status,
        # Cooldown tracking fields
        'cooldown_remaining': campaign.cooldown_remaining,
        'cooldown_status': campaign.cooldown_status,
        'current_batch': campaign.current_batch,
        'total_batches': campaign.total_batches,
        'use_advanced_controls': campaign.use_advanced_controls
    })


@login_required
@require_POST
def stop_campaign(request, campaign_id):
    """
    Stop a running or pending campaign by setting status to 'paused'.
    The background task will detect this and halt gracefully.
    """
    campaign = get_object_or_404(WASenderCampaign, id=campaign_id, user=request.user)
    if campaign.status in ['running', 'pending']:
        campaign.status = 'paused'
        campaign.save(update_fields=['status'])
        from django.contrib import messages
        messages.success(request, f"Campaign '{campaign.name}' has been stopped.")
    else:
        from django.contrib import messages
        messages.info(request, f"Campaign '{campaign.name}' is not running.")
    # Support AJAX stop as well
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'status': campaign.status})
    return redirect('whatsappapi:campaign_detail', campaign_id=campaign_id)


# ==================== Retry Failed Messages ====================

@login_required
@require_POST
def retry_failed_messages(request, campaign_id):
    """
    Retry all failed messages for a given campaign using the campaign's
    session and message configuration. Does not change UI; returns to
    campaign detail with a summary.
    """
    campaign = get_object_or_404(WASenderCampaign, id=campaign_id, user=request.user)

    # Base failed set for this session
    failed_qs = WASenderMessage.objects.filter(
        session=campaign.session,
        status__in=['failed', 'queued']
    ).order_by('created_at')

    # Prefer metadata linkage when available
    linked_failed = failed_qs.filter(metadata__campaign_id=campaign.id)
    if linked_failed.exists():
        failed_qs = linked_failed
    else:
        # Fallback to campaign window and recipients
        failed_qs = failed_qs.filter(created_at__gte=campaign.created_at)
        if campaign.recipients:
            phones = [r.get('phone') for r in campaign.recipients if r.get('phone')]
            if phones:
                failed_qs = failed_qs.filter(recipient__in=list(set(phones)))

    total_to_retry = failed_qs.count()
    if total_to_retry == 0:
        messages.info(request, "No failed messages to retry for this campaign.")
        return redirect('whatsappapi:campaign_detail', campaign_id=campaign.id)

    service = WASenderService()
    sent_now = 0
    failed_again = 0

    for fm in failed_qs:
        try:
            recipient = fm.recipient
            if (campaign.message_type or 'text') == 'text':
                result = service.send_text_message(campaign.session, recipient, campaign.message_template)
            else:
                media_url = campaign.media_url or campaign.attachment_url
                caption = campaign.description or ''
                result = service.send_media_message(
                    campaign.session,
                    recipient,
                    media_url,
                    campaign.message_type,
                    caption=caption
                )
            if result and result.status == 'sent':
                sent_now += 1
            else:
                failed_again += 1
        except Exception:
            failed_again += 1

    # Refresh campaign stats
    try:
        campaign.update_stats()
    except Exception:
        pass

    messages.info(
        request,
        f"Retried {total_to_retry} failed messages: {sent_now} sent, {failed_again} still failed."
    )
    return redirect('whatsappapi:campaign_detail', campaign_id=campaign.id)


# ==================== Retry Single Recipient ====================

@login_required
@require_POST
def retry_single_recipient(request, campaign_id):
    """
    Retry sending a campaign message to a single recipient.
    Expects 'recipient' in POST body; uses campaign configuration.
    """
    campaign = get_object_or_404(WASenderCampaign, id=campaign_id, user=request.user)
    recipient = (request.POST.get('recipient') or '').strip()

    if not recipient:
        messages.error(request, "Recipient phone is required to retry.")
        return redirect('whatsappapi:campaign_detail', campaign_id=campaign.id)

    service = WASenderService()
    sent_now = False
    try:
        if (campaign.message_type or 'text') == 'text':
            msg = service.send_text_message(campaign.session, recipient, campaign.message_template)
        else:
            media_url = campaign.media_url or campaign.attachment_url
            caption = campaign.description or ''
            msg = service.send_media_message(
                campaign.session,
                recipient,
                media_url,
                campaign.message_type,
                caption=caption
            )

        # Tag message with campaign id for linkage
        if msg:
            try:
                meta = msg.metadata or {}
                meta.update({'campaign_id': campaign.id})
                msg.metadata = meta
                msg.save(update_fields=['metadata'])
            except Exception:
                pass
            sent_now = (msg.status == 'sent')
    except Exception:
        sent_now = False

    # Refresh stats
    try:
        campaign.update_stats()
    except Exception:
        pass

    # Don't add Django messages here - let the campaign detail page show stats instead
    return redirect('whatsappapi:campaign_detail', campaign_id=campaign.id)


# ==================== Contacts Management ====================

@login_required
@never_cache
def contacts(request):
    """
    Contacts management page
    List all uploaded contact lists
    """
    from .models import ContactList
    
    # Get all contact lists for current user
    contact_lists = ContactList.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'contact_lists': contact_lists,
    }
    return render(request, 'whatsappapi/contacts.html', context)


@login_required
def upload_contacts(request):
    """
    Upload and process contact list file (CSV/XLSX)
    """
    from django.http import JsonResponse
    from .models import ContactList, Contact
    import pandas as pd
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get uploaded file and parameters
        uploaded_file = request.FILES.get('file')
        list_name = request.POST.get('list_name')
        country_code = request.POST.get('country_code', '')
        country_type = request.POST.get('country_type', 'single')
        
        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        
        if not list_name:
            return JsonResponse({'success': False, 'error': 'Contact list name is required'})
        
        # Read file based on extension
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            df = pd.read_csv(uploaded_file)
        elif file_extension in ['xlsx', 'xls']:
            df = pd.read_excel(uploaded_file)
        else:
            return JsonResponse({'success': False, 'error': 'Unsupported file format'})
        
        # Validate required columns
        if 'phone' not in df.columns:
            return JsonResponse({'success': False, 'error': 'Phone column is required in the file'})
        
        # Get all CSV headers (excluding 'phone') for dynamic fields
        available_fields = [col for col in df.columns if col != 'phone']
        
        # Create ContactList
        contact_list = ContactList.objects.create(
            user=request.user,
            name=list_name,
            file_name=uploaded_file.name,
            country_code=f"+{country_code}" if country_code else '',
            total_contacts=0,  # Will update after processing
            valid_contacts=0,
            available_fields=available_fields  # Store CSV headers dynamically
        )

        # Broadcast: new contact list created
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            from django.utils import timezone

            channel_layer = get_channel_layer()
            formatted_created_at = timezone.localtime(contact_list.created_at).strftime("%d/%m/%Y, %H:%M:%S")
            async_to_sync(channel_layer.group_send)(
                f"user_updates_{request.user.id}",
                {
                    'type': 'contact_list_created',
                    'list': {
                        'id': contact_list.id,
                        'name': contact_list.name,
                        'total_contacts': contact_list.total_contacts,
                        'created_at': formatted_created_at,
                    }
                }
            )
        except Exception:
            pass
        
        # Remove duplicate phone numbers from DataFrame
        df_cleaned = df.drop_duplicates(subset=['phone'], keep='first')
        duplicates_removed = len(df) - len(df_cleaned)
        
        if duplicates_removed > 0:
            logger.info(f"Removed {duplicates_removed} duplicate phone numbers from upload")
        
        # Process contacts (limit to 10000)
        contacts_added = 0
        seen_phones = set()  # Track phones to prevent duplicates within this upload
        
        for index, row in df_cleaned.head(10000).iterrows():
            try:
                phone = str(row.get('phone', '')).strip()
                
                if not phone:
                    continue
                
                # Add country code if single country mode
                if country_type == 'single' and country_code:
                    # Remove leading zeros
                    phone = phone.lstrip('0')
                    # Add country code if not present
                    if not phone.startswith(country_code):
                        phone = f"{country_code}{phone}"
                
                # Skip if this phone was already processed in this upload
                if phone in seen_phones:
                    logger.debug(f"Skipping duplicate phone in same upload: {phone}")
                    continue
                
                # Check if phone already exists in THIS contact list
                if Contact.objects.filter(contact_list=contact_list, phone_number=phone).exists():
                    logger.debug(f"Skipping existing phone in contact list: {phone}")
                    continue
                
                seen_phones.add(phone)
                
                # Build dynamic fields dictionary (all columns except 'phone')
                contact_fields = {}
                for col in df.columns:
                    if col != 'phone':
                        value = row.get(col, '')
                        contact_fields[col] = str(value) if pd.notna(value) else ''
                
                # Create contact with dynamic fields
                Contact.objects.create(
                    contact_list=contact_list,
                    phone_number=phone,
                    fields=contact_fields,  # Store all CSV fields dynamically
                    # Backward compatibility - populate specific fields if they exist
                    first_name=contact_fields.get('first_name', ''),
                    last_name=contact_fields.get('last_name', ''),
                    email=contact_fields.get('email', ''),
                    custom_field_1=contact_fields.get('custom_field_1', ''),
                    custom_field_2=contact_fields.get('custom_field_2', ''),
                    custom_field_3=contact_fields.get('custom_field_3', '')
                )
                contacts_added += 1
                
            except Exception as e:
                logger.warning(f"Error processing contact at row {index}: {e}")
                continue
        
        # Update contact list counts
        contact_list.total_contacts = contacts_added
        contact_list.save()

        # Broadcast: contacts count updated
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            from django.utils import timezone

            channel_layer = get_channel_layer()
            formatted_created_at = timezone.localtime(contact_list.created_at).strftime("%d/%m/%Y, %H:%M:%S")
            async_to_sync(channel_layer.group_send)(
                f"user_updates_{request.user.id}",
                {
                    'type': 'contacts_count_updated',
                    'list': {
                        'id': contact_list.id,
                        'name': contact_list.name,
                        'total_contacts': contact_list.total_contacts,
                        'created_at': formatted_created_at,
                    }
                }
            )
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'total_contacts': contacts_added,
            'list_id': contact_list.id,
            'duplicates_removed': duplicates_removed,
            'message': f'Successfully imported {contacts_added} contacts' + (f' ({duplicates_removed} duplicates removed)' if duplicates_removed > 0 else '')
        })
        
    except Exception as e:
        logger.error(f"Error uploading contacts: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def delete_contact_list(request, list_id):
    """
    Delete a contact list and all its contacts
    """
    from django.http import JsonResponse
    from .models import ContactList
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get the contact list
        contact_list = ContactList.objects.filter(id=list_id, user=request.user).first()
        
        if not contact_list:
            return JsonResponse({'success': False, 'error': 'Contact list not found'})
        
        # Delete the list (cascades to contacts)
        contact_list.delete()

        # Broadcast: contact list deleted
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync

            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_updates_{request.user.id}",
                {
                    'type': 'contact_list_deleted',
                    'list_id': int(list_id),
                }
            )
        except Exception:
            pass

        return JsonResponse({'success': True, 'message': 'Contact list deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_contact_list_fields(request, list_id):
    """
    AJAX endpoint to get available fields from a contact list
    """
    from django.http import JsonResponse
    from .models import ContactList
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        contact_list = ContactList.objects.filter(id=list_id, user=request.user).first()
        
        if not contact_list:
            return JsonResponse({'success': False, 'error': 'Contact list not found'})
        
        # Return available fields from the contact list
        available_fields = contact_list.available_fields if contact_list.available_fields else []
        
        return JsonResponse({
            'success': True,
            'fields': available_fields,
            'list_name': contact_list.name,
            'total_contacts': contact_list.total_contacts
        })
        
    except Exception as e:
        logger.error(f"Error fetching contact list fields: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def check_moderation(request):
    """
    AJAX endpoint to check message content for moderation issues without sending
    Returns: {"blocked": bool, "risk_score": int, "reasons": [...], "illegal_words": [...], "word_positions": {...}}
    """
    from django.http import JsonResponse
    from .moderation import evaluate_content
    import json
    import re
    
    try:
        # Parse request body
        try:
            if request.content_type and 'application/json' in request.content_type:
                payload = json.loads(request.body or '{}')
            else:
                payload = request.POST
        except Exception:
            payload = request.POST
        
        message = (payload.get('message') or '').strip()
        
        if not message:
            return JsonResponse({'error': 'Message cannot be empty'}, status=400)
        
        # Run moderation check
        moderation = evaluate_content(message)
        
        # Extract policy hits for illegal words highlighting
        illegal_words = []
        word_positions = {}  # Track positions of illegal words in message
        
        # Extract keywords from reasons
        for reason in moderation.get('reasons', []):
            reason_str = str(reason).lower()
            # Extract category labels from reasons like "ai:violence" or "classifier:drug_trade"
            if ':' in reason_str:
                keyword = reason_str.split(':')[-1].replace('_', ' ')
                if keyword not in illegal_words and keyword.strip():
                    illegal_words.append(keyword)
        
        # Also extract from policy_hits if available
        for hit in moderation.get('policy_hits', []):
            category = hit.get('category', '').replace('_', ' ').lower()
            if category and category not in illegal_words:
                illegal_words.append(category)
        
        # Find positions of illegal keywords in the message for highlighting
        message_lower = message.lower()
        for word in illegal_words:
            positions = []
            start = 0
            while True:
                pos = message_lower.find(word, start)
                if pos == -1:
                    break
                # Find word boundaries
                end = pos + len(word)
                positions.append({'start': pos, 'end': end})
                start = end
            if positions:
                word_positions[word] = positions
        
        # Determine if blocked based on AI_ONLY_GATE setting
        import os
        from django.conf import settings
        _ai_only_val = os.environ.get('AI_ONLY_GATE', None)
        if _ai_only_val is None:
            try:
                _ai_only_val = getattr(settings, 'AI_ONLY_GATE', 'true')
            except Exception:
                _ai_only_val = 'true'
        ai_only_gate = str(_ai_only_val).lower() in {'1', 'true', 'yes'}
        
        is_blocked = moderation.get('blocked') or (ai_only_gate and moderation.get('requires_review'))
        
        return JsonResponse({
            'blocked': is_blocked,
            'requires_review': moderation.get('requires_review', False),
            'risk_score': int(moderation.get('risk_score', 0)),
            'reasons': moderation.get('reasons', []),
            'illegal_words': illegal_words,
            'word_positions': word_positions,
            'message': 'Content moderation check completed'
        })
        
    except Exception as e:
        logger.error(f"Error checking moderation: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def rename_contact_list(request, list_id):
    """
    Rename a contact list
    """
    from django.http import JsonResponse
    from .models import ContactList
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        data = json.loads(request.body)
        new_name = data.get('name', '').strip()
        
        if not new_name:
            return JsonResponse({'success': False, 'error': 'Name is required'})
        
        # Get the contact list
        contact_list = ContactList.objects.filter(id=list_id, user=request.user).first()
        
        if not contact_list:
            return JsonResponse({'success': False, 'error': 'Contact list not found'})
        
        # Update the name
        contact_list.name = new_name
        contact_list.save()

        # Broadcast: contact list renamed
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            from django.utils import timezone

            channel_layer = get_channel_layer()
            formatted_created_at = timezone.localtime(contact_list.created_at).strftime("%d/%m/%Y, %H:%M:%S")
            async_to_sync(channel_layer.group_send)(
                f"user_updates_{request.user.id}",
                {
                    'type': 'contact_list_renamed',
                    'list': {
                        'id': contact_list.id,
                        'name': contact_list.name,
                        'total_contacts': contact_list.total_contacts,
                        'created_at': formatted_created_at,
                    }
                }
            )
        except Exception:
            pass

        return JsonResponse({'success': True, 'message': 'Contact list renamed successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def export_contact_list_csv(request, list_id):
    """
    Export contact list as CSV
    """
    from django.http import HttpResponse
    from .models import ContactList
    import csv
    
    try:
        contact_list = ContactList.objects.filter(id=list_id, user=request.user).first()
        
        if not contact_list:
            return HttpResponse('Contact list not found', status=404)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{contact_list.name}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['phone', 'first_name', 'last_name', 'email', 'custom_field_1', 'custom_field_2', 'custom_field_3'])
        
        for contact in contact_list.contacts.all():
            writer.writerow([
                contact.phone_number,
                contact.first_name or '',
                contact.last_name or '',
                contact.email or '',
                contact.custom_field_1 or '',
                contact.custom_field_2 or '',
                contact.custom_field_3 or ''
            ])
        
        return response
        
    except Exception as e:
        return HttpResponse(f'Error: {str(e)}', status=500)


@login_required
def export_contact_list_excel(request, list_id):
    """
    Export contact list as Excel
    """
    from django.http import HttpResponse
    from .models import ContactList
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    try:
        contact_list = ContactList.objects.filter(id=list_id, user=request.user).first()
        
        if not contact_list:
            return HttpResponse('Contact list not found', status=404)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = contact_list.name[:31]  # Excel limit
        
        # Headers
        headers = ['phone', 'first_name', 'last_name', 'email', 'custom_field_1', 'custom_field_2', 'custom_field_3']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add contacts
        for contact in contact_list.contacts.all():
            ws.append([
                contact.phone_number,
                contact.first_name or '',
                contact.last_name or '',
                contact.email or '',
                contact.custom_field_1 or '',
                contact.custom_field_2 or '',
                contact.custom_field_3 or ''
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{contact_list.name}.xlsx"'
        
        return response
        
    except Exception as e:
        return HttpResponse(f'Error: {str(e)}', status=500)


def download_sample_csv(request):
    """
    Download sample CSV template for contact upload
    """
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_contacts.csv"'
    
    writer = csv.writer(response)
    # Headers with phone as first field
    writer.writerow(['phone', 'first_name', 'last_name', 'email', 'custom_field_1', 'custom_field_2', 'custom_field_3'])
    writer.writerow(['919876543210', 'John', 'Doe', 'john@example.com', 'Company A', 'Premium', 'Delhi'])
    writer.writerow(['918765432109', 'Jane', 'Smith', 'jane@example.com', 'Company B', 'Standard', 'Mumbai'])
    writer.writerow(['917654321098', 'Mike', 'Johnson', 'mike@example.com', 'Company C', 'Premium', 'Bangalore'])
    writer.writerow(['916543210987', 'Sarah', 'Williams', 'sarah@example.com', 'Company D', 'Basic', 'Pune'])
    writer.writerow(['915432109876', 'David', 'Brown', 'david@example.com', 'Company E', 'Premium', 'Hyderabad'])
    
    return response


def download_sample_excel(request):
    """
    Download sample Excel template for contact upload
    """
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    # Add headers with phone as first field
    headers = ['phone', 'first_name', 'last_name', 'email', 'custom_field_1', 'custom_field_2', 'custom_field_3']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add sample data
    sample_data = [
        ['919876543210', 'John', 'Doe', 'john@example.com', 'Company A', 'Premium', 'Delhi'],
        ['918765432109', 'Jane', 'Smith', 'jane@example.com', 'Company B', 'Standard', 'Mumbai'],
        ['917654321098', 'Mike', 'Johnson', 'mike@example.com', 'Company C', 'Premium', 'Bangalore'],
        ['916543210987', 'Sarah', 'Williams', 'sarah@example.com', 'Company D', 'Basic', 'Pune'],
        ['915432109876', 'David', 'Brown', 'david@example.com', 'Company E', 'Premium', 'Hyderabad']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18  # phone
    ws.column_dimensions['B'].width = 15  # first_name
    ws.column_dimensions['C'].width = 15  # last_name
    ws.column_dimensions['D'].width = 25  # email
    ws.column_dimensions['E'].width = 15  # custom_field_1
    ws.column_dimensions['F'].width = 15  # custom_field_2
    ws.column_dimensions['G'].width = 15  # custom_field_3
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_contacts.xlsx"'
    
    return response


@login_required
def export_campaign_messages_excel(request, campaign_id):
    """
    Export campaign messages as Excel file
    """
    from django.http import HttpResponse
    from django.utils import timezone
    import json
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    try:
        campaign = get_object_or_404(WASenderCampaign, id=campaign_id, user=request.user)
        
        # Base queryset: session messages
        messages_qs = WASenderMessage.objects.filter(
            session=campaign.session
        ).order_by('-created_at')
        
        # Prefer metadata linkage (campaign_id) if available; use JSON key transform
        linked_qs = messages_qs.filter(metadata__campaign_id=campaign.id)
        if linked_qs.exists():
            messages = linked_qs
        else:
            # Fallback to created_at window
            messages = messages_qs.filter(created_at__gte=campaign.created_at)
        
        # Further filter by recipients only if not using metadata linkage
        if not linked_qs.exists():
            if campaign.recipients:
                phone_numbers = [recipient.get('phone') for recipient in campaign.recipients if recipient.get('phone')]
                if phone_numbers:
                    unique_phone_numbers = list(set(phone_numbers))
                    messages = messages.filter(recipient__in=unique_phone_numbers)
            else:
                # SQLite does not support DISTINCT ON; pick latest per recipient via Subquery
                try:
                    from django.db.models import OuterRef, Subquery
                    latest_ids = WASenderMessage.objects.filter(
                        session=campaign.session,
                        recipient=OuterRef('recipient'),
                        created_at__gte=campaign.created_at
                    ).order_by('-created_at').values('id')[:1]
                    messages = messages.filter(id=Subquery(latest_ids)).order_by('-created_at')
                except Exception:
                    messages = messages.order_by('-created_at')
        
        wb = openpyxl.Workbook()
        
        # ==================== SUMMARY SHEET ====================
        ws_summary = wb.active
        ws_summary.title = "Campaign Summary"
        
        # Calculate statistics
        total_messages = messages.count()
        messages_sent = messages.filter(status__in=['sent', 'delivered', 'read']).count()
        messages_delivered = messages.filter(status__in=['delivered', 'read']).count()
        messages_read = messages.filter(status='read').count()
        messages_failed = messages.filter(status='failed').count()
        messages_pending = messages.filter(status='pending').count()
        
        success_rate = 0
        if campaign.total_recipients > 0:
            success_rate = round((messages_delivered / campaign.total_recipients) * 100, 2)
        
        # Title
        ws_summary.append(['CAMPAIGN REPORT'])
        ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_summary.merge_cells('A1:B1')
        
        ws_summary.append([])  # Empty row
        
        # Campaign Info
        ws_summary.append(['Campaign Name:', campaign.name])
        ws_summary.append(['Created:', campaign.created_at.strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append(['Status:', campaign.status.upper()])
        ws_summary.append(['WhatsApp Number:', campaign.session.connected_phone_number or campaign.session.phone_number])
        
        ws_summary.append([])  # Empty row
        
        # Statistics Header
        ws_summary.append(['CAMPAIGN STATISTICS'])
        ws_summary['A8'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_summary['A8'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_summary.merge_cells('A8:B8')
        
        ws_summary.append([])  # Empty row
        
        # Statistics
        ws_summary.append(['Total Recipients:', campaign.total_recipients])
        ws_summary.append(['Messages Sent:', messages_sent])
        ws_summary.append(['✅ Delivered:', messages_delivered])
        ws_summary.append(['📖 Read:', messages_read])
        ws_summary.append(['❌ Failed:', messages_failed])
        ws_summary.append(['⏳ Pending:', messages_pending])
        ws_summary.append([])
        ws_summary.append(['Success Rate:', f"{success_rate}%"])
        
        # Style statistics
        for row in range(10, 18):
            ws_summary[f'A{row}'].font = Font(bold=True)
            if row == 17:  # Success Rate
                ws_summary[f'A{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                ws_summary[f'B{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                ws_summary[f'B{row}'].font = Font(bold=True, size=14, color="70AD47")
        
        # Adjust column widths for summary
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        
        # ==================== MESSAGES DETAIL SHEET ====================
        ws = wb.create_sheet(title="Message Details")
        
        # Headers
        headers = ['Phone', 'Status', 'Sent At', 'Message ID']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add messages
        for message in messages:
            # Extract phone number from recipient field
            phone = message.recipient
            
            # Try to parse JSON format to get phone number
            if message.recipient.startswith('{'):
                try:
                    recipient_data = json.loads(message.recipient)
                    phone = recipient_data.get('phone', message.recipient)
                except:
                    pass  # If JSON parsing fails, use recipient as is
            
            # Color code status
            row_data = [
                phone,
                message.status.upper(),
                message.sent_at.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S') if message.sent_at else '',
                message.message_id or ''
            ]
            ws.append(row_data)
            
            # Apply color based on status
            current_row = ws.max_row
            if message.status == 'delivered' or message.status == 'read':
                fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif message.status == 'failed':
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif message.status == 'pending':
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                fill = None
            
            if fill:
                for cell in ws[current_row]:
                    cell.fill = fill
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{campaign.name}_messages.xlsx"'
        
        return response
        
    except Exception as e:
        return HttpResponse(f'Error: {str(e)}', status=500)

@login_required
def save_draft(request):
    """
    Save campaign as draft template
    """
    from django.http import JsonResponse
    from .models import CampaignTemplate, ContactList
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        campaign_name = request.POST.get('campaign_name', '').strip()
        message = request.POST.get('message', '').strip()
        contact_list_id = request.POST.get('contact_list_id')
        draft_id = request.POST.get('draft_id')
        attachment = request.FILES.get('attachment')
        
        if not campaign_name:
            return JsonResponse({'success': False, 'error': 'Campaign name is required'})
        
        # Validate campaign name format (only lowercase, numbers, underscores)
        import re
        if not re.match(r'^[a-z0-9_]+$', campaign_name):
            return JsonResponse({'success': False, 'error': 'Campaign name must contain only lowercase letters, numbers, and underscores (e.g., new_year_offer_2025)'})
        
        if not message:
            return JsonResponse({'success': False, 'error': 'Message is required'})
        
        # Get contact list
        contact_list = None
        if contact_list_id:
            contact_list = ContactList.objects.filter(id=contact_list_id, user=request.user).first()
        
        # Update existing draft or create new
        if draft_id:
            draft = CampaignTemplate.objects.filter(id=draft_id, user=request.user).first()
            if draft:
                draft.name = campaign_name
                draft.message = message
                draft.contact_list = contact_list
                if attachment:
                    draft.attachment = attachment
                    draft.attachment_type = get_file_type(attachment.name)
                draft.save()
            else:
                return JsonResponse({'success': False, 'error': 'Draft not found'})
        else:
            draft = CampaignTemplate.objects.create(
                user=request.user,
                name=campaign_name,
                message=message,
                contact_list=contact_list,
                attachment=attachment,
                attachment_type=get_file_type(attachment.name) if attachment else None,
                status='draft'
            )
        
        return JsonResponse({
            'success': True,
            'draft_id': draft.id,
            'message': 'Draft saved successfully'
        })
        
    except Exception as e:
        logger.error(f"Error saving draft: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_draft(request, draft_id):
    """
    Get draft template details
    """
    from django.http import JsonResponse
    from .models import CampaignTemplate
    
    try:
        draft = CampaignTemplate.objects.filter(id=draft_id, user=request.user).first()
        
        if not draft:
            return JsonResponse({'success': False, 'error': 'Draft not found'})
        
        return JsonResponse({
            'success': True,
            'id': draft.id,
            'name': draft.name,
            'message': draft.message,
            'contact_list_id': draft.contact_list.id if draft.contact_list else None,
            'contact_list_name': draft.contact_list.name if draft.contact_list else None,
            'attachment': draft.attachment.url if draft.attachment else None,
            'attachment_type': draft.attachment_type
        })
        
    except Exception as e:
        logger.error(f"Error getting draft: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def delete_draft(request, draft_id):
    """
    Delete draft template
    """
    from django.http import JsonResponse
    from .models import CampaignTemplate
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        draft = CampaignTemplate.objects.filter(id=draft_id, user=request.user).first()
        
        if not draft:
            return JsonResponse({'success': False, 'error': 'Draft not found'})
        
        draft.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Draft deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting draft: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


def get_file_type(filename):
    """
    Determine file type from filename
    """
    ext = filename.lower().split('.')[-1]
    if ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
        return 'image'
    elif ext in ['mp4', 'avi', 'mov', 'mkv']:
        return 'video'
    elif ext in ['pdf', 'doc', 'docx', 'txt', 'xls', 'xlsx']:
        return 'document'
    return 'other'


@login_required
def drafts_list(request):
    """
    Display all saved draft templates
    """
    from .models import CampaignTemplate
    
    drafts = CampaignTemplate.objects.filter(user=request.user, status='draft').order_by('-updated_at')
    
    context = {
        'drafts': drafts,
    }
    
    return render(request, 'whatsappapi/drafts.html', context)


@login_required
def drafts_list(request):
    """
    Display all saved draft templates
    """
    from .models import CampaignTemplate
    
    drafts = CampaignTemplate.objects.filter(user=request.user, status='draft').order_by('-updated_at')
    
    context = {
        'drafts': drafts,
    }
    
    return render(request, 'whatsappapi/drafts.html', context)
